
# <--------------------- For anynoe not familiar with python just change the file_path variable and faculty_file_path
# and run command "python3 daily.py in mac" or "python daily.py in windows  -------------------->


#importing librarires 

import pandas as pd
import re
import openpyxl
from openpyxl import load_workbook
from openpyxl.styles import Font
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import FormulaRule

file_path = "reg-capstat-2025-spring-11-14-2024.xlsx"

# wb = openpyxl.load_workbook(file_path)
# status = wb.active.cell(wb.active.min_row, 1).value 
# print(status)
# wb.active.delete_rows(wb.active.min_row, 1) 
# wb.save(file_path)

faculty_file_path = "faculty_dept.xlsx"

df = pd.read_excel(file_path)
df1 = pd.read_excel(faculty_file_path)

# creation of file_a filter with sse ssw ise ide em sys etc

def create_df_a(df):
    df = df[['Course', 'Instructor(s)/Teaching Assistant', 'Title', 
             'Minimum Units', 'Section Capacity', 
             'Enrollment Count', 'Meeting Patterns', 'Building/Room']]
    
    if 'Course' not in df.columns:
        raise KeyError("'Course' column is missing from the DataFrame.")
    
    df['Course'] = df['Course'].astype(str)
    df = df[df['Course'].str.contains('AAI|EM|SYS|SSW|ISE|IDE', case=False)]
    df = df[~df['Course'].str.contains('EMT')]
    if df.empty:
        print("No courses found with the specified prefixes.")
        return None, df

    df['Section'] = df['Course'].str.extract(r'-(.*?)(?:/|$)')
    df['Meeting Patterns'] = df['Meeting Patterns'].fillna('')
    df['Building/Room'] = df['Building/Room'].fillna('')
    df['Title'] = df['Title'].str.strip()

    final_output = df.sort_values(['Course', 'Instructor(s)/Teaching Assistant'])
    output_file = 'combined_courses_file_a.xlsx'
    final_output.to_excel(output_file, index=False)
    
    return output_file, final_output



output_file_combined, final_df = create_df_a(df)
print(final_df.shape)

# creation of file_b filter with of faculties with se/ se a

df = pd.read_excel(file_path)
df1 = pd.read_excel(faculty_file_path)

def create_df_b(df, faculty_dept):
    se_sea_faculty = faculty_dept[faculty_dept['Department Name'].isin(['SE', 'SE-A'])]
    df = df[df['Instructor(s)/Teaching Assistant'].isin(se_sea_faculty['Instructor(s)/Teaching Assistant'])]
    df = df[['Course', 'Instructor(s)/Teaching Assistant', 'Title', 
             'Minimum Units', 'Section Capacity', 
             'Enrollment Count', 'Meeting Patterns', 'Building/Room']]
    
    df['Meeting Patterns'] = df['Meeting Patterns'].fillna('')
    df['Building/Room'] = df['Building/Room'].fillna('')
    df['Title'] = df['Title'].str.strip()
    
    output_file = 'combined_courses_file_b.xlsx'
    df.to_excel(output_file, index=False)
    
    return output_file, df

output_file_combined, final_df = create_df_b(df, df1)

print(final_df.shape)


#file _c

 # creation of file_c filter with file_a - file_b
file_a = pd.read_excel('combined_courses_file_a.xlsx')
file_b = pd.read_excel('combined_courses_file_b.xlsx')

key_columns = ['Course', 'Instructor(s)/Teaching Assistant', 'Title', 
             'Minimum Units', 'Section Capacity', 
             'Enrollment Count', 'Meeting Patterns', 'Building/Room']

file_c = file_b.merge(file_a[key_columns], on=key_columns, how='left', indicator=True)
file_c = file_c[file_c['_merge'] == 'left_only'].drop(columns='_merge')

file_c.to_excel('combined_courses_file_c.xlsx', index=False)
print(file_c.shape)


# # creation of file_d filter with file_a + file_c

file_d = pd.concat([file_a, file_c], ignore_index=True)
print('Before duplicates count',file_d.shape)

duplicates = file_d[file_d.duplicated()]
print(duplicates)
file_d.drop_duplicates(inplace=True)
print('After dropping duplicates count',file_d.shape)

result = pd.merge(file_d, df1, on='Instructor(s)/Teaching Assistant', how='left')

result.to_excel('combined_courses_file_d.xlsx', index=False)

def clean_course_number(title: str) -> str:
    course_numbers = title.split('/')
    cleaned_numbers = [re.sub(r'-[^/]+$', '', num) for num in course_numbers]
    return '/'.join(cleaned_numbers)


def extract_section(title: str) -> str:
    match = re.search(r'-([^/]+)', title)
    return match.group(1) if match else ''

#pivot

def add_pivot_tables_to_existing_excel(output_file, df):

    def create_web_campus_sheet(sheet_name):

        web_df = df[
            (df['Department Name'].isin(['SE', 'SE-A'])) &
            (df['Building/Room'].str.contains('WebCampus', na=False))
        ]

        web_output = pd.DataFrame(columns=['Instructor(s)/Teaching Assistant', 'Combined_Course', 'Building/Room', 'SE', 'SE-A', 'Grand Total'])

        for instructor, group in web_df.groupby('Instructor(s)/Teaching Assistant'):
            se_total = group[group['Department Name'] == 'SE']['Enrollment Count'].sum()
            se_a_total = group[group['Department Name'] == 'SE-A']['Enrollment Count'].sum()

            web_output = pd.concat([web_output, pd.DataFrame({
                'Instructor(s)/Teaching Assistant': [instructor],
                'Combined_Course': [''],
                'Building/Room': [''],
                'SE': [se_total],
                'SE-A': [se_a_total],
                'Grand Total': [se_total + se_a_total]
            })], ignore_index=True)

            for _, row in group.iterrows():
                web_output = pd.concat([web_output, pd.DataFrame({
                    'Instructor(s)/Teaching Assistant': [''],
                    'Combined_Course': [row['Course']],
                    'Building/Room': [row['Building/Room']],
                    'SE': [row['Enrollment Count'] if row['Department Name'] == 'SE' else 0],
                    'SE-A': [row['Enrollment Count'] if row['Department Name'] == 'SE-A' else 0],
                    'Grand Total': [row['Enrollment Count']]
                })], ignore_index=True)

        grand_total_se = web_output['SE'].sum()
        grand_total_se_a = web_output['SE-A'].sum()
        web_output = pd.concat([web_output, pd.DataFrame({
            'Instructor(s)/Teaching Assistant': ['Grand Total'],
            'Combined_Course': [''],
            'Building/Room': [''],
            'SE': [grand_total_se],
            'SE-A': [grand_total_se_a],
            'Grand Total': [grand_total_se + grand_total_se_a]
        })], ignore_index=True)

        web_output.to_excel(writer, sheet_name=sheet_name, index=False)
        
    def create_courses_sheet(sheet_name):
        df['Course Title'] = df['Course'].apply(clean_course_number)
        df['Section'] = df['Course'].apply(extract_section)
        df['Instructor'] = df['Instructor(s)/Teaching Assistant']

        web_df = df[(df['Department Name'].isin(['SE', 'SE-A']))]

        web_output = pd.DataFrame(columns=['Course', 'Title', 'Section', 'Instructor(s)/Teaching Assistant', 'Building/Room', 'SE', 'SE-A', 'Grand Total'])
        for instructor, group in web_df.groupby('Course Title'):

            se_total = group[group['Department Name'] == 'SE']['Enrollment Count'].sum()
            se_a_total = group[group['Department Name'] == 'SE-A']['Enrollment Count'].sum()
            sections_present = group['Section'].unique()
            should_combine = {'A', 'A-U'}.issubset(sections_present)
            course_row = {
                'Course': instructor,
                'Title': '',  
                'Section': '',
                'Instructor(s)/Teaching Assistant': '',
                'Building/Room': '',
                'SE': se_total,
                'SE-A': se_a_total,
                'Grand Total': se_total + se_a_total
            }
            web_output = pd.concat([web_output, pd.DataFrame([course_row])], ignore_index=True)

            if should_combine:
                combined_entry = {
                    'Course': '',
                    'Title': group['Title'].iloc[0],
                    'Section': 'A & A-U',
                    'Instructor(s)/Teaching Assistant': group['Instructor(s)/Teaching Assistant'].iloc[0],
                    'Building/Room': group['Building/Room'].iloc[0],
                    'SE': group[group['Section'].isin(['A', 'A-U']) & (group['Department Name'] == 'SE')]['Enrollment Count'].sum(),
                    'SE-A': group[group['Section'].isin(['A', 'A-U']) & (group['Department Name'] == 'SE-A')]['Enrollment Count'].sum(),
                    'Grand Total': (group[group['Section'].isin(['A', 'A-U']) & (group['Department Name'] == 'SE')]['Enrollment Count'].sum() +
                                group[group['Section'].isin(['A', 'A-U']) & (group['Department Name'] == 'SE-A')]['Enrollment Count'].sum())
                }
                web_output = pd.concat([web_output, pd.DataFrame([combined_entry])], ignore_index=True)

            for _, row in group.iterrows():
                if not should_combine or row['Section'] not in ['A', 'A-U']:
                    web_output = pd.concat([web_output, pd.DataFrame({
                        'Course': '',
                        'Title': [row['Title']], 
                        'Section': [row['Section']],
                        'Instructor(s)/Teaching Assistant': [row['Instructor(s)/Teaching Assistant']],
                        'Building/Room': [row['Building/Room']],
                        'SE': [row['Enrollment Count'] if row['Department Name'] == 'SE' else 0],
                        'SE-A': [row['Enrollment Count'] if row['Department Name'] == 'SE-A' else 0],
                        'Grand Total': [row['Enrollment Count']]
                    })], ignore_index=True)

        grand_total_se = web_output['SE'].sum()
        grand_total_se_a = web_output['SE-A'].sum()
        web_output = pd.concat([web_output, pd.DataFrame({
            'Course': ['Grand Total'],
            'Title': [''],
            'Section': [''],
            'Instructor(s)/Teaching Assistant': [''],
            'Building/Room': [''],
            'SE': [grand_total_se],
            'SE-A': [grand_total_se_a],
            'Grand Total': [grand_total_se + grand_total_se_a]
        })], ignore_index=True)

        web_output.to_excel(writer, sheet_name=sheet_name, index=False)


    with pd.ExcelWriter(output_file, engine='openpyxl', mode='a') as writer:
        create_web_campus_sheet('Web Campus')
        create_courses_sheet('Course Offerings')

    wb = load_workbook(output_file)

    bold_font = Font(bold=True)

    web_ws = wb['Web Campus']

    for cell in web_ws[1]:
        cell.font = bold_font

    for row in web_ws.iter_rows(min_row=2, max_row=web_ws.max_row, min_col=1, max_col=1):
        if row[0].value:
            row[0].font = bold_font

    wb.save(output_file)


    course_ws = wb['Course Offerings']

    for cell in course_ws[1]:
        cell.font = bold_font

    for row in course_ws.iter_rows(min_row=2, max_row=course_ws.max_row, min_col=1, max_col=1):
        if row[0].value:
            row[0].font = bold_font

    wb.save(output_file)


add_pivot_tables_to_existing_excel('combined_courses_file_d.xlsx', result)

result1 = result[result['Department Name'] == 'SE']
result2 = result[result['Department Name'] == 'SE-A']
result3 = result[(result['Department Name'] != 'SE') & (result['Department Name'] != 'SE-A')]

def process_course_load(data: pd.DataFrame, default_load_status='On Load') -> pd.DataFrame:
    df = data.copy()
    df['Course Title'] = df['Course'].apply(clean_course_number)
    df['Section_Only'] = df['Course'].apply(extract_section)
    df['Instructor'] = df['Instructor(s)/Teaching Assistant']

    course_sections = df.groupby(['Instructor', 'Course Title'])['Section_Only'].apply(list).reset_index()
    def should_combine_sections(sections):
        return set(sections) == {'A', 'A-U'}

    course_sections['should_combine'] = course_sections['Section_Only'].apply(should_combine_sections)
    combine_map = dict(zip(
        zip(course_sections['Instructor'], course_sections['Course Title']), 
        course_sections['should_combine']
    ))

    grouped = df.groupby(['Instructor', 'Course Title', 'Section_Only']).agg({
        'Title': 'first',
        'Enrollment Count': 'sum',
        'Department Name': 'first',
        'Stream': 'first',
        'Minimum Units': 'sum'
    }).reset_index()
    total_enrollment = grouped.groupby('Instructor')['Enrollment Count'].sum().reset_index()
    total_enrollment.rename(columns={'Enrollment Count': 'Total Enrollment'}, inplace=True)
    
    total_minimum_units = grouped.groupby('Instructor')['Minimum Units'].sum().reset_index()
    total_minimum_units.rename(columns={'Minimum Units': 'Total Minimum Units'}, inplace=True)

    output_rows = []
    prev_instructor = None
    for instructor, instructor_group in grouped.groupby('Instructor'):
        if prev_instructor is not None and prev_instructor != instructor:
            output_rows.append({
                'Instructor': '',
                'Course Title': '',
                'Sections': '',
                'Title': '',
                'Enrollment Count': '',
                'Department Name': '',
                'Stream': '',
                'Load Status': '',
                'Payment Initiator': '',
                'Minimum Units': '',
            })

        total_row = {
            'Instructor': instructor,
            'Course Title': '',
            'Sections': '',
            'Title': '',
            'Enrollment Count': total_enrollment[total_enrollment['Instructor'] == instructor]['Total Enrollment'].iloc[0],
            'Department Name': '',
            'Stream': '',
            'Load Status': default_load_status,
            'Payment Initiator': 'N/A',
            'Minimum Units': total_minimum_units[total_minimum_units['Instructor'] == instructor]['Total Minimum Units'].iloc[0]
        }
        output_rows.append(total_row)

        processed_courses = set()
        for _, row in instructor_group.iterrows():
            course_key = (row['Instructor'], row['Course Title'])

            if course_key in processed_courses:
                continue

            if combine_map.get(course_key, False):
                course_data = instructor_group[
                    (instructor_group['Course Title'] == row['Course Title'])
                ]

                combined_row = {
                    'Instructor': '',
                    'Course Title': row['Course Title'],
                    'Sections': 'A & A-U',
                    'Title': row['Title'],
                    'Enrollment Count': course_data['Enrollment Count'].sum(),
                    'Department Name': row['Department Name'],
                    'Stream': row['Stream'],
                    'Load Status': default_load_status,  
                    'Payment Initiator': 'N/A',
                    'Minimum Units': course_data['Minimum Units'].sum(),
                }
                output_rows.append(combined_row)
                processed_courses.add(course_key)
            else:
                course_row = {
                    'Instructor': '',
                    'Course Title': row['Course Title'],
                    'Sections': row['Section_Only'],
                    'Title': row['Title'],
                    'Enrollment Count': row['Enrollment Count'],
                    'Department Name': row['Department Name'],
                    'Stream': row['Stream'],
                    'Load Status': default_load_status,  
                    'Payment Initiator': 'N/A',
                    'Minimum Units': row['Minimum Units'],
                }
                output_rows.append(course_row)
                processed_courses.add(course_key)

        prev_instructor = instructor

    result = pd.DataFrame(output_rows)
    return result[[
        'Instructor',
        'Course Title',
        'Sections',
        'Title',
        'Enrollment Count',
        'Minimum Units',
        'Department Name',
        'Stream',
        'Load Status',
        'Payment Initiator',
    ]]


def save_to_existing_excel_with_dropdown_and_color(df: pd.DataFrame, existing_file: str, sheet_name: str):
    with pd.ExcelWriter(existing_file, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
        df.to_excel(writer, sheet_name=sheet_name, index=False)

        workbook = writer.book
        worksheet = writer.sheets[sheet_name]

        for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row):
            load_status_cell = row[df.columns.get_loc('Load Status')]
            payment_initiator_cell = row[df.columns.get_loc('Payment Initiator')]

            if not load_status_cell.value:
                load_status_cell.value = 'N/A'
            if not payment_initiator_cell.value:
                payment_initiator_cell.value = 'N/A'

        for row in range(2, worksheet.max_row + 1):
            instructor_cell = worksheet.cell(row=row, column=1)
            if instructor_cell.value:
                instructor_cell.font = Font(bold=True)

        load_status_validation = DataValidation(
            type="list",
            formula1='"N/A,Online,On Load,Extra Pay"',
            allow_blank=True
        )
        payment_validation = DataValidation(
            type="list",
            formula1='"N/A,SE,Corporate,WebCampus"',
            allow_blank=True
        )

        worksheet.add_data_validation(load_status_validation)
        worksheet.add_data_validation(payment_validation)

        load_status_col_idx = df.columns.get_loc('Load Status') + 1
        payment_initiator_col_idx = df.columns.get_loc('Payment Initiator') + 1

        load_status_col = get_column_letter(load_status_col_idx)
        payment_initiator_col = get_column_letter(payment_initiator_col_idx)

        for row in range(2, worksheet.max_row + 1):
            load_status_validation.add(f'{load_status_col}{row}')
            payment_validation.add(f'{payment_initiator_col}{row}')

        colors = {
            "Online": "FFEB9C",
            "On Load": "9BC2E6",
            "Extra Pay": "E2EFDA"
        }

        for status, color in colors.items():
            formula = f'${load_status_col}2="{status}"'
            rule = FormulaRule(
                formula=[formula],
                stopIfTrue=True,
                fill=PatternFill(start_color=color, end_color=color, fill_type='solid')
            )
            full_range = f'A2:{get_column_letter(worksheet.max_column)}{worksheet.max_row}'
            worksheet.conditional_formatting.add(full_range, rule)

        for col in range(1, len(df.columns) + 1):
            max_length = 0
            column = get_column_letter(col)
            for cell in worksheet[column]:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            worksheet.column_dimensions[column].width = max_length + 2

processed_df = process_course_load(result)
save_to_existing_excel_with_dropdown_and_color(processed_df, 'combined_courses_file_d.xlsx', 'Faculty Assignments')
processed_df1 = process_course_load(result1)
save_to_existing_excel_with_dropdown_and_color(processed_df1, 'combined_courses_file_d.xlsx', 'Faculty Assignments-SE')
processed_df2 = process_course_load(result2, default_load_status='N/A')
save_to_existing_excel_with_dropdown_and_color(processed_df2, 'combined_courses_file_d.xlsx', 'Faculty Assignments-SE-A')
processed_df3 = process_course_load(result3, default_load_status='N/A')
save_to_existing_excel_with_dropdown_and_color(processed_df3, 'combined_courses_file_d.xlsx', 'Faculty Assignments-Other')


def format_excel(file_path):
    wb = load_workbook(file_path)
    yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    bold_font = Font(bold=True)
    center_alignment = Alignment(horizontal="center", vertical="center")

    for sheet in wb.sheetnames:
        ws = wb[sheet]
        for col in ws.columns:
            max_length = 0
            for cell in col:
                cell.alignment = center_alignment
                if cell.row == 1: 
                    cell.font = bold_font
                    cell.fill = yellow_fill
                max_length = max(max_length, len(str(cell.value) if cell.value else ""))
            ws.column_dimensions[col[0].column_letter].width = max_length + 2

    wb.save(file_path)

format_excel('combined_courses_file_d.xlsx')



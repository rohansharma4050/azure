
# <--------------------- For anynoe not familiar with python just change the file_path variable and faculty_file_path
# and run command "python3 daily.py in mac" or "python daily.py in windows  -------------------->


#importing librarires 

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font
from openpyxl.styles import Font, Alignment, PatternFill

file_path = "Reg-Cap_stat_11-11-24.xlsx"
faculty_file_path = "faculty_dept.xlsx"

df = pd.read_excel(file_path)
df1 = pd.read_excel(faculty_file_path)

# creation of file_a filter with sse ssw ise ide em sys etc

def create_df_a(df):

    df = df[['Course', 'Instructor(s)/Teaching Assistant', 'Maximum Units', 'Section Capacity',
             'Title', 'Enrollment Count', 'Meeting Patterns', 'Building/Room']]
    df = df[df['Course'].str.contains('AAI|EM|SYS|SSW|ISE|IDE', case=False)]

    if df.empty:
        print("No courses found with the specified prefixes.")
        return None, df
    df['Course_split'] = df['Course'].str.split('/')
    df = df.explode('Course_split')
    df = df.drop(columns=['Course']).rename(columns={'Course_split': 'Course'})
    df = df[df['Course'].str.contains('AAI|EM|SYS|SSW|ISE|IDE', case=False)]

    if df.empty:
        print("No courses found after splitting and filtering.")
        return None, df
    df['Department'] = df['Course'].str.extract(r'([A-Z]+)')[0]
    df['Course_Number'] = df['Course'].str.extract(r'(\d+)')[0]
    df['Section'] = df['Course'].str.extract(r'-([A-Z0-9]+)')[0]
    df['Meeting Patterns'] = df['Meeting Patterns'].fillna('')
    df['Building/Room'] = df['Building/Room'].fillna('')
    section_count_df = df.groupby(['Course_Number', 'Instructor(s)/Teaching Assistant']).size().reset_index(name='Section Count')
    combined_courses = df.groupby(['Course_Number', 'Instructor(s)/Teaching Assistant']).agg({
        'Department': lambda x: '_'.join(sorted(set(x))),
        'Section': lambda x: '/'.join(sorted(set(x))), 
        'Section Capacity': lambda x: '/'.join(x.astype(str)),
        'Enrollment Count': lambda x: '+'.join(x.astype(str)),
        'Maximum Units': lambda x: ', '.join(x.astype(str).unique()),
        'Meeting Patterns': lambda x: '/'.join(x.unique()),
        'Building/Room': lambda x: '/'.join(x.unique())
    }).reset_index()

    combined_courses = combined_courses.merge(section_count_df, on=['Course_Number', 'Instructor(s)/Teaching Assistant'], how='left')
    combined_courses['Combined_Course'] = combined_courses.apply(
        lambda row: f"{row['Department']}_{row['Course_Number']}", axis=1
    )
    combined_courses['Total Enrollment Count'] = combined_courses['Enrollment Count'].apply(
        lambda x: sum(map(int, x.split('+')))
    )
    final_output = combined_courses[['Combined_Course', 'Section Count', 'Section', 'Section Capacity',
                                     'Enrollment Count', 'Total Enrollment Count', 'Meeting Patterns',
                                     'Building/Room', 'Instructor(s)/Teaching Assistant']]
    final_output = final_output.sort_values(['Combined_Course', 'Instructor(s)/Teaching Assistant'])
    output_file = 'combined_courses_instructors.xlsx'
    final_output.to_excel(output_file, index=False)

    return output_file, final_output
output_file_combined, final_df = create_df_a(df)
final_df = final_df.merge(df1, how='left', on='Instructor(s)/Teaching Assistant')
final_df.to_excel('combined_courses_file_a.xlsx', index=False)

print(final_df.shape)



# creation of file_b filter with of faculties with se/ se a

df = pd.read_excel(file_path)
df1 = pd.read_excel(faculty_file_path)


def create_df_b(df, faculty_dept):
    se_sea_faculty = faculty_dept[faculty_dept['Department Name'].isin(['SE', 'SE-A'])]
    df = df[df['Instructor(s)/Teaching Assistant'].isin(se_sea_faculty['Instructor(s)/Teaching Assistant'])]
    df = df[['Course', 'Instructor(s)/Teaching Assistant', 'Maximum Units', 'Section Capacity',
             'Title', 'Enrollment Count', 'Meeting Patterns', 'Building/Room']]
    df['Course_split'] = df['Course'].str.split('/')
    df = df.explode('Course_split').drop(columns=['Course']).rename(columns={'Course_split': 'Course'})
    df['Department'] = df['Course'].str.extract(r'([A-Z]+)')[0]
    df['Course_Number'] = df['Course'].str.extract(r'(\d+)')[0]
    df['Section'] = df['Course'].str.extract(r'-([A-Z0-9]+)')[0]
    df['Meeting Patterns'] = df['Meeting Patterns'].fillna('')
    df['Building/Room'] = df['Building/Room'].fillna('')
    section_count_df = df.groupby(['Course_Number', 'Instructor(s)/Teaching Assistant']).size().reset_index(name='Section Count')
    combined_courses = df.groupby(['Course_Number', 'Instructor(s)/Teaching Assistant']).agg({
        'Department': lambda x: '_'.join(sorted(set(x))),
        'Section': lambda x: '/'.join(sorted(set(x))),
        'Section Capacity': lambda x: '/'.join(x.astype(str)),
        'Enrollment Count': lambda x: '+'.join(x.astype(str)),
        'Maximum Units': lambda x: ', '.join(x.astype(str).unique()),
        'Meeting Patterns': lambda x: '/'.join(x.unique()),
        'Building/Room': lambda x: '/'.join(x.unique())
    }).reset_index()
    combined_courses = combined_courses.merge(section_count_df, on=['Course_Number', 'Instructor(s)/Teaching Assistant'], how='left')
    combined_courses = combined_courses.merge(
        se_sea_faculty[['Instructor(s)/Teaching Assistant', 'Department Name']],
        on='Instructor(s)/Teaching Assistant',
        how='left'
    )
    combined_courses['Combined_Course'] = combined_courses.apply(
        lambda row: f"{row['Department']}_{row['Course_Number']}", axis=1
    )
    combined_courses['Total Enrollment Count'] = combined_courses['Enrollment Count'].apply(
        lambda x: sum(map(int, x.split('+')))
    )
    final_output = combined_courses[['Combined_Course', 'Section Count', 'Section', 'Section Capacity',
                                     'Enrollment Count', 'Total Enrollment Count', 'Meeting Patterns',
                                     'Building/Room', 'Instructor(s)/Teaching Assistant', 'Department Name']]
    final_output = final_output.sort_values(['Combined_Course', 'Department Name', 'Instructor(s)/Teaching Assistant'])
    output_file = 'combined_courses_file_b.xlsx'
    final_output.to_excel(output_file, index=False)
    return output_file, final_output

output_file_combined, final_df = create_df_b(df, df1)

print(final_df.shape)



# creation of file_c filter with file_a - file_b
file_a = pd.read_excel('combined_courses_file_a.xlsx')
file_b = pd.read_excel('combined_courses_file_b.xlsx')

key_columns = ["Combined_Course", "Section Count", "Section", "Section Capacity",
               "Enrollment Count", "Total Enrollment Count", "Meeting Patterns",
               "Building/Room", "Instructor(s)/Teaching Assistant"]

file_c = file_b.merge(file_a[key_columns], on=key_columns, how='left', indicator=True)
file_c = file_c[file_c['_merge'] == 'left_only'].drop(columns='_merge')

file_c.to_excel('combined_courses_file_c.xlsx', index=False)
print(file_c.shape)




# # creation of file_d filter with file_a + file_c

file_d = pd.concat([file_a, file_c], ignore_index=True)
file_d.to_excel('combined_courses_file_d.xlsx', index=False)

print(file_d.shape)



# pivot table code on base file with formatting

def add_pivot_tables_to_existing_excel(output_file, df):
    def create_department_sheet(department_name, sheet_name):
        dept_df = df[df['Department Name'] == department_name]
        if dept_df.empty:
            print(f"No data found for {department_name}.")
            return
        dept_pivot = dept_df.pivot_table(
            values='Total Enrollment Count',
            index=['Instructor(s)/Teaching Assistant', 'Combined_Course'],
            aggfunc='sum'
        ).reset_index()

        formatted_pivot = pd.DataFrame(columns=['Instructor(s)/Teaching Assistant', 'Combined_Course', 'Total Enrollment Count'])

        for instructor, group in dept_pivot.groupby('Instructor(s)/Teaching Assistant'):
            instructor_total = group['Total Enrollment Count'].sum()
            formatted_pivot = pd.concat([formatted_pivot, pd.DataFrame({
                'Instructor(s)/Teaching Assistant': [instructor],
                'Combined_Course': [''],
                'Total Enrollment Count': [instructor_total]
            })], ignore_index=True)
            for _, row in group.iterrows():
                formatted_pivot = pd.concat([formatted_pivot, pd.DataFrame({
                    'Instructor(s)/Teaching Assistant': [''],
                    'Combined_Course': [row['Combined_Course']],
                    'Total Enrollment Count': [row['Total Enrollment Count']]
                })], ignore_index=True)

        grand_total = formatted_pivot[formatted_pivot['Combined_Course'] == '']['Total Enrollment Count'].sum()
        formatted_pivot = pd.concat([formatted_pivot, pd.DataFrame({
            'Instructor(s)/Teaching Assistant': ['Grand Total'],
            'Combined_Course': [''],
            'Total Enrollment Count': [grand_total]
        })], ignore_index=True)

        formatted_pivot.to_excel(writer, sheet_name=sheet_name, index=False)

    def create_online_sheet(sheet_name):
        online_df = df[df['Department Name'].isin(['SE', 'SE-A'])]
        online_pivot = online_df.pivot_table(
            values='Section Count',
            index=['Instructor(s)/Teaching Assistant', 'Building/Room', 'Combined_Course'],
            columns='Department Name',
            aggfunc='sum',
            fill_value=0
        ).reset_index()

        formatted_online_pivot = pd.DataFrame(columns=['Instructor(s)/Teaching Assistant', 'Combined_Course', 'Building/Room', 'SE', 'SE-A', 'Grand Total'])

        for instructor, group in online_pivot.groupby('Instructor(s)/Teaching Assistant'):
            instructor_total_se = group['SE'].sum()
            instructor_total_se_a = group['SE-A'].sum()
            grand_total = instructor_total_se + instructor_total_se_a

            formatted_online_pivot = pd.concat([formatted_online_pivot, pd.DataFrame({
                'Instructor(s)/Teaching Assistant': [instructor],
                'Combined_Course': [''],
                'Building/Room': [''],
                'SE': [instructor_total_se],
                'SE-A': [instructor_total_se_a],
                'Grand Total': [grand_total]
            })], ignore_index=True)

            for _, row in group.iterrows():
                formatted_online_pivot = pd.concat([formatted_online_pivot, pd.DataFrame({
                    'Instructor(s)/Teaching Assistant': [''],
                    'Combined_Course': [row['Combined_Course']],
                    'Building/Room': [row['Building/Room']],
                    'SE': [row['SE']],
                    'SE-A': [row['SE-A']],
                    'Grand Total': ['']
                })], ignore_index=True)

        grand_total_se = formatted_online_pivot[formatted_online_pivot['Combined_Course'] == '']['SE'].sum()
        grand_total_se_a = formatted_online_pivot[formatted_online_pivot['Combined_Course'] == '']['SE-A'].sum()
        overall_grand_total = grand_total_se + grand_total_se_a

        formatted_online_pivot = pd.concat([formatted_online_pivot, pd.DataFrame({
            'Instructor(s)/Teaching Assistant': ['Grand Total'],
            'Combined_Course': [''],
            'Building/Room': [''],
            'SE': [grand_total_se],
            'SE-A': [grand_total_se_a],
            'Grand Total': [overall_grand_total]
        })], ignore_index=True)

        formatted_online_pivot.to_excel(writer, sheet_name=sheet_name, index=False)

    with pd.ExcelWriter(output_file, engine='openpyxl', mode='a') as writer:
        create_department_sheet('SE-A', 'LOADS-SE-A')
        create_department_sheet('SE', 'LOADS-SE')
        create_online_sheet('Online')

    wb = load_workbook(output_file)

    for sheet_name in ['LOADS-SE-A', 'LOADS-SE']:
        ws = wb[sheet_name]

        bold_font = Font(bold=True)
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=3):
            if row[0].value and row[0].value != '':
                row[0].font = bold_font

        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, values_only=False):
            if len(row) > 2:
                if row[0].value is not None and row[1].value == '':
                    row[0].font = bold_font

        last_row = ws.max_row
        ws.cell(row=last_row, column=1).font = bold_font
        ws.cell(row=last_row, column=3).font = bold_font

    online_ws = wb['Online']
    bold_font = Font(bold=True)
    for cell in online_ws[1]:
        cell.font = bold_font

    for row in online_ws.iter_rows(min_row=2, max_row=online_ws.max_row, min_col=1, max_col=1):
        if row[0].value:
            row[0].font = bold_font

    wb.save(output_file)

add_pivot_tables_to_existing_excel('combined_courses_file_d.xlsx', file_d)

def format_excel(file_path):
    wb = load_workbook(file_path)
    yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    bold_font = Font(bold=True)
    center_alignment = Alignment(horizontal="center", vertical="center")

    for sheet in wb.sheetnames:
        ws = wb[sheet]
        for col in ws.columns:
            max_length = 0
            for cell in col:
                cell.alignment = center_alignment
                if cell.row == 1: 
                    cell.font = bold_font
                    cell.fill = yellow_fill
                max_length = max(max_length, len(str(cell.value) if cell.value else ""))
            ws.column_dimensions[col[0].column_letter].width = max_length + 2

    wb.save(file_path)

format_excel('combined_courses_file_d.xlsx')


print('<-------------Code Successfully Done -- Download combined_courses_file_d.xlsx NOW----------------->')
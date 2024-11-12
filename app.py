import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font
from azure.storage.blob import BlobServiceClient
import io
from datetime import datetime

# Azure Storage connection details
CONNECTION_STRING = 
CONTAINER_NAME = "output"
INPUT_CONTAINER_NAME = "input"
blob_service_client = BlobServiceClient.from_connection_string(CONNECTION_STRING)

def read_excel_from_azure(blob_name):
    blob_client = blob_service_client.get_blob_client(container=INPUT_CONTAINER_NAME, blob=blob_name)
    stream = io.BytesIO()
    blob_client.download_blob().readinto(stream)
    stream.seek(0)
    return pd.read_excel(stream)

def write_excel_to_azure(df, blob_name):
    stream = io.BytesIO()
    with pd.ExcelWriter(stream, engine='openpyxl') as writer:
        df.to_excel(writer, index=False)
    stream.seek(0)
    blob_client = blob_service_client.get_blob_client(container=CONTAINER_NAME, blob=blob_name)
    blob_client.upload_blob(stream, overwrite=True)

# Load initial data files from Azure
df = read_excel_from_azure("Reg-Cap_stat_11-11-24.xlsx")
df1 = read_excel_from_azure("faculty_dept.xlsx")

# Process data to create file_a
def create_df_a(df):
    df = df[['Course', 'Instructor(s)/Teaching Assistant', 'Maximum Units', 'Section Capacity',
             'Title', 'Enrollment Count', 'Meeting Patterns', 'Building/Room']]
    df = df[df['Course'].str.contains('AAI|EM|SYS|SSW|ISE|IDE', case=False)]

    if df.empty:
        print("No courses found with the specified prefixes.")
        return None, df

    df['Course_split'] = df['Course'].str.split('/')
    df = df.explode('Course_split')
    df = df.drop(columns=['Course']).rename(columns={'Course_split': 'Course'})
    df = df[df['Course'].str.contains('AAI|EM|SYS|SSW|ISE|IDE', case=False)]

    if df.empty:
        print("No courses found after splitting and filtering.")
        return None, df

    df['Department'] = df['Course'].str.extract(r'([A-Z]+)')[0]
    df['Course_Number'] = df['Course'].str.extract(r'(\d+)')[0]
    df['Section'] = df['Course'].str.extract(r'-([A-Z0-9]+)')[0]
    df['Meeting Patterns'] = df['Meeting Patterns'].fillna('')
    df['Building/Room'] = df['Building/Room'].fillna('')

    section_count_df = df.groupby(['Course_Number', 'Instructor(s)/Teaching Assistant']).size().reset_index(name='Section Count')
    combined_courses = df.groupby(['Course_Number', 'Instructor(s)/Teaching Assistant']).agg({
        'Department': lambda x: '_'.join(sorted(set(x))),
        'Section': lambda x: '/'.join(sorted(set(x))),
        'Section Capacity': lambda x: '/'.join(x.astype(str)),
        'Enrollment Count': lambda x: '+'.join(x.astype(str)),
        'Maximum Units': lambda x: ', '.join(x.astype(str).unique()),
        'Meeting Patterns': lambda x: '/'.join(x.unique()),
        'Building/Room': lambda x: '/'.join(x.unique())
    }).reset_index()

    combined_courses = combined_courses.merge(section_count_df, on=['Course_Number', 'Instructor(s)/Teaching Assistant'], how='left')
    combined_courses['Combined_Course'] = combined_courses.apply(lambda row: f"{row['Department']}_{row['Course_Number']}", axis=1)
    combined_courses['Total Enrollment Count'] = combined_courses['Enrollment Count'].apply(lambda x: sum(map(int, x.split('+'))))

    final_output = combined_courses[['Combined_Course', 'Section Count', 'Section', 'Section Capacity',
                                     'Enrollment Count', 'Total Enrollment Count', 'Meeting Patterns',
                                     'Building/Room', 'Instructor(s)/Teaching Assistant']]

    final_output = final_output.sort_values(['Combined_Course', 'Instructor(s)/Teaching Assistant'])

    output_file_a = f"combined_courses_file_a_{datetime.today().strftime('%Y%m%d')}.xlsx"
    write_excel_to_azure(final_output, output_file_a)
    return output_file_a, final_output

output_file_a, final_df_a = create_df_a(df)

df = read_excel_from_azure("Reg-Cap_stat_11-11-24.xlsx")
df1 = read_excel_from_azure("faculty_dept.xlsx")
# Process data to create file_b
def create_df_b(df, faculty_dept):
    se_sea_faculty = faculty_dept[faculty_dept['Department Name'].isin(['SE', 'SE-A'])]
    df = df[df['Instructor(s)/Teaching Assistant'].isin(se_sea_faculty['Instructor(s)/Teaching Assistant'])]

    df = df[['Course', 'Instructor(s)/Teaching Assistant', 'Maximum Units', 'Section Capacity',
             'Title', 'Enrollment Count', 'Meeting Patterns', 'Building/Room']]

    df['Course_split'] = df['Course'].str.split('/')
    df = df.explode('Course_split')
    df = df.drop(columns=['Course']).rename(columns={'Course_split': 'Course'})

    df['Department'] = df['Course'].str.extract(r'([A-Z]+)')[0]
    df['Course_Number'] = df['Course'].str.extract(r'(\d+)')[0]
    df['Section'] = df['Course'].str.extract(r'-([A-Z0-9]+)')[0]
    df['Meeting Patterns'] = df['Meeting Patterns'].fillna('')
    df['Building/Room'] = df['Building/Room'].fillna('')

    section_count_df = df.groupby(['Course_Number', 'Instructor(s)/Teaching Assistant']).size().reset_index(name='Section Count')
    combined_courses = df.groupby(['Course_Number', 'Instructor(s)/Teaching Assistant']).agg({
        'Department': lambda x: '_'.join(sorted(set(x))),
        'Section': lambda x: '/'.join(sorted(set(x))),
        'Section Capacity': lambda x: '/'.join(x.astype(str)),
        'Enrollment Count': lambda x: '+'.join(x.astype(str)),
        'Maximum Units': lambda x: ', '.join(x.astype(str).unique()),
        'Meeting Patterns': lambda x: '/'.join(x.unique()),
        'Building/Room': lambda x: '/'.join(x.unique())
    }).reset_index()

    combined_courses = combined_courses.merge(section_count_df, on=['Course_Number', 'Instructor(s)/Teaching Assistant'], how='left')
    combined_courses = combined_courses.merge(se_sea_faculty[['Instructor(s)/Teaching Assistant', 'Department Name']],
                                              on='Instructor(s)/Teaching Assistant', how='left')

    combined_courses['Combined_Course'] = combined_courses.apply(lambda row: f"{row['Department']}_{row['Course_Number']}", axis=1)
    combined_courses['Total Enrollment Count'] = combined_courses['Enrollment Count'].apply(lambda x: sum(map(int, x.split('+'))))

    final_output = combined_courses[['Combined_Course', 'Section Count', 'Section', 'Section Capacity',
                                     'Enrollment Count', 'Total Enrollment Count', 'Meeting Patterns',
                                     'Building/Room', 'Instructor(s)/Teaching Assistant', 'Department Name']]

    final_output = final_output.sort_values(['Combined_Course', 'Department Name', 'Instructor(s)/Teaching Assistant'])

    output_file_b = f"combined_courses_file_b_{datetime.today().strftime('%Y%m%d')}.xlsx"
    write_excel_to_azure(final_output, output_file_b)
    return output_file_b, final_output

output_file_b, final_df_b = create_df_b(df, df1)

# Generate file_c and file_d
def create_file_c_and_d(file_a, file_b):
    key_columns = ["Combined_Course", "Section Count", "Section", "Section Capacity",
                   "Enrollment Count", "Total Enrollment Count", "Meeting Patterns",
                   "Building/Room", "Instructor(s)/Teaching Assistant"]

    file_c = file_b.merge(file_a[key_columns], on=key_columns, how='left', indicator=True)
    file_c = file_c[file_c['_merge'] == 'left_only'].drop(columns='_merge')

    output_file_c = f"combined_courses_file_c_{datetime.today().strftime('%Y%m%d')}.xlsx"
    write_excel_to_azure(file_c, output_file_c)

    file_d = pd.concat([file_a, file_c], ignore_index=True)

    output_file_d = f"combined_courses_file_d_{datetime.today().strftime('%Y%m%d')}.xlsx"
    write_excel_to_azure(file_d, output_file_d)
    return output_file_c, output_file_d, file_d

output_file_c, output_file_d, final_df_d = create_file_c_and_d(final_df_a, final_df_b)

# Function to add pivot tables and save to Azure
import pandas as pd
import io
from azure.storage.blob import BlobServiceClient
# Your existing function to add pivot tables
def add_pivot_tables_to_existing_excel(output_file, df):
    def create_department_sheet(department_name, sheet_name):
        dept_df = df[df['Department Name'] == department_name]
        if dept_df.empty:
            print(f"No data found for {department_name}.")
            return
        dept_pivot = dept_df.pivot_table(
            values='Total Enrollment Count',
            index=['Instructor(s)/Teaching Assistant', 'Combined_Course'],
            aggfunc='sum'
        ).reset_index()

        formatted_pivot = pd.DataFrame(columns=['Instructor(s)/Teaching Assistant', 'Combined_Course', 'Total Enrollment Count'])

        for instructor, group in dept_pivot.groupby('Instructor(s)/Teaching Assistant'):
            instructor_total = group['Total Enrollment Count'].sum()
            formatted_pivot = pd.concat([formatted_pivot, pd.DataFrame({
                'Instructor(s)/Teaching Assistant': [instructor],
                'Combined_Course': [''],
                'Total Enrollment Count': [instructor_total]
            })], ignore_index=True)
            for _, row in group.iterrows():
                formatted_pivot = pd.concat([formatted_pivot, pd.DataFrame({
                    'Instructor(s)/Teaching Assistant': [''],
                    'Combined_Course': [row['Combined_Course']],
                    'Total Enrollment Count': [row['Total Enrollment Count']]
                })], ignore_index=True)

        grand_total = formatted_pivot[formatted_pivot['Combined_Course'] == '']['Total Enrollment Count'].sum()
        formatted_pivot = pd.concat([formatted_pivot, pd.DataFrame({
            'Instructor(s)/Teaching Assistant': ['Grand Total'],
            'Combined_Course': [''],
            'Total Enrollment Count': [grand_total]
        })], ignore_index=True)

        formatted_pivot.to_excel(writer, sheet_name=sheet_name, index=False)

    def create_online_sheet(sheet_name):
        online_df = df[df['Department Name'].isin(['SE', 'SE-A'])]
        online_pivot = online_df.pivot_table(
            values='Section Count',
            index=['Instructor(s)/Teaching Assistant', 'Building/Room', 'Combined_Course'],
            columns='Department Name',
            aggfunc='sum',
            fill_value=0
        ).reset_index()

        formatted_online_pivot = pd.DataFrame(columns=['Instructor(s)/Teaching Assistant', 'Combined_Course', 'Building/Room', 'SE', 'SE-A', 'Grand Total'])

        for instructor, group in online_pivot.groupby('Instructor(s)/Teaching Assistant'):
            instructor_total_se = group['SE'].sum()
            instructor_total_se_a = group['SE-A'].sum()
            grand_total = instructor_total_se + instructor_total_se_a

            formatted_online_pivot = pd.concat([formatted_online_pivot, pd.DataFrame({
                'Instructor(s)/Teaching Assistant': [instructor],
                'Combined_Course': [''],
                'Building/Room': [''],
                'SE': [instructor_total_se],
                'SE-A': [instructor_total_se_a],
                'Grand Total': [grand_total]
            })], ignore_index=True)

            for _, row in group.iterrows():
                formatted_online_pivot = pd.concat([formatted_online_pivot, pd.DataFrame({
                    'Instructor(s)/Teaching Assistant': [''],
                    'Combined_Course': [row['Combined_Course']],
                    'Building/Room': [row['Building/Room']],
                    'SE': [row['SE']],
                    'SE-A': [row['SE-A']],
                    'Grand Total': ['']
                })], ignore_index=True)

        grand_total_se = formatted_online_pivot[formatted_online_pivot['Combined_Course'] == '']['SE'].sum()
        grand_total_se_a = formatted_online_pivot[formatted_online_pivot['Combined_Course'] == '']['SE-A'].sum()
        overall_grand_total = grand_total_se + grand_total_se_a

        formatted_online_pivot = pd.concat([formatted_online_pivot, pd.DataFrame({
            'Instructor(s)/Teaching Assistant': ['Grand Total'],
            'Combined_Course': [''],
            'Building/Room': [''],
            'SE': [grand_total_se],
            'SE-A': [grand_total_se_a],
            'Grand Total': [overall_grand_total]
        })], ignore_index=True)

        formatted_online_pivot.to_excel(writer, sheet_name=sheet_name, index=False)

    # Check if the file exists in Azure Storage
    blob_client = blob_service_client.get_blob_client(container=CONTAINER_NAME, blob=output_file)
    stream = io.BytesIO()

    try:
        # Try to download the file if it exists
        blob_client.download_blob().readinto(stream)
        stream.seek(0)
        mode = 'a'  # Append if the file exists
    except Exception as e:
        print("File not found in Azure. Creating a new file.")
        mode = 'w'  # Write if the file doesn't exist

    # Write pivot tables to the Excel file
    with pd.ExcelWriter(stream, engine='openpyxl', mode=mode) as writer:
        create_department_sheet('SE-A', 'LOADS-SE-A')
        create_department_sheet('SE', 'LOADS-SE')
        create_online_sheet('Online')

    # Apply formatting
    stream.seek(0)  # Reset stream position after writing
    wb = load_workbook(stream)
    for sheet_name in ['LOADS-SE-A', 'LOADS-SE']:
        ws = wb[sheet_name]

        bold_font = Font(bold=True)
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=3):
            if row[0].value and row[0].value != '':
                row[0].font = bold_font

        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, values_only=False):
            if len(row) > 2:
                if row[0].value is not None and row[1].value == '':
                    row[0].font = bold_font

        last_row = ws.max_row
        ws.cell(row=last_row, column=1).font = bold_font
        ws.cell(row=last_row, column=3).font = bold_font

    online_ws = wb['Online']
    bold_font = Font(bold=True)
    for cell in online_ws[1]:
        cell.font = bold_font

    for row in online_ws.iter_rows(min_row=2, max_row=online_ws.max_row, min_col=1, max_col=1):
        if row[0].value:
            row[0].font = bold_font

    wb.save(stream)

    # Upload the final file to Azure Blob Storage
    stream.seek(0)
    blob_client.upload_blob(stream, overwrite=True)
    print(f"File uploaded to Azure Blob Storage as {output_file}")

# Add pivot tables using the combined DataFrame (file_d)
add_pivot_tables_to_existing_excel( output_file_d,final_df_d)


